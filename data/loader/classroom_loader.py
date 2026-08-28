from pathlib import Path
from openpyxl import load_workbook
from engine.models.classroom import Classroom


class ClassroomLoader:
    def __init__(self, classroom_file: str):
        self.classroom_file = Path(classroom_file)

    def load(self) -> list[Classroom]:
        if not self.classroom_file.exists():
            raise FileNotFoundError(f"Classroom file not found: {self.classroom_file}")

        workbook = load_workbook(self.classroom_file, data_only=True)
        worksheet = workbook.active

        classrooms: list[Classroom] = []

        def parse_int(val, default: int) -> int:
            if val is None:
                return default
            val_str = str(val).strip().lower()
            if val_str in ("", "nan", "none"):
                return default
            try:
                # Convert float strings like '6.0' cleanly to int
                return int(float(val_str))
            except (ValueError, TypeError):
                return default

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None or str(row[0]).strip() == "":
                continue

            room_no = str(row[0]).strip()
            rows = parse_int(row[1] if len(row) > 1 else None, default=5)
            benches = parse_int(row[2] if len(row) > 2 else None, default=3)
            seats = parse_int(row[3] if len(row) > 3 else None, default=3)

            classrooms.append(
                Classroom(
                    room_no=room_no,
                    rows=rows,
                    benches_per_row=benches,
                    seats_per_bench=seats,
                )
            )

        workbook.close()
        return classrooms


def load_classrooms(classroom_file: str) -> list[Classroom]:
    return ClassroomLoader(classroom_file).load()