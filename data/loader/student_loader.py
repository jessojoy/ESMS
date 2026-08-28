from pathlib import Path
from openpyxl import load_workbook
from engine.models.student import Student


class StudentLoader:
    def __init__(self, student_file: str, timetable_file: str):
        self.student_file = Path(student_file)
        self.timetable_file = Path(timetable_file)

    def load(self) -> list[Student]:
        timetable = self._load_timetable()
        workbook = load_workbook(self.student_file, data_only=True)
        students: list[Student] = []

        for sheet_name in workbook.sheetnames:
            if sheet_name.lower() == "master overview":
                continue

            worksheet = workbook[sheet_name]
            department = self._get_department(sheet_name)
            semester = self._get_semester(sheet_name)
            section = self._get_section(sheet_name)

            exams = timetable.get((department, semester), [])
            if not exams:
                continue

            for row in worksheet.iter_rows(min_row=7, values_only=True):
                if row[0] is None:
                    break

                roll_no = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                base_reg_no = str(row[3]).strip()
                name = str(row[4]).strip()

                # Disambiguate registration numbers for multi-section classes
                reg_no = f"{base_reg_no}-{section}" if section else base_reg_no

                for exam in exams:
                    code = str(exam["subject_code"] or "").strip()
                    subj_name = str(exam["subject_name"] or "").strip()

                    students.append(
                        Student(
                            register_no=reg_no,
                            name=name,
                            department=department,
                            semester=semester,
                            section=section,
                            subject_code=code if code and code.lower() != "nan" else subj_name,
                            subject_name=subj_name,
                            exam_date=str(exam["exam_date"] or "").strip(),
                            session=str(exam["session"] or "").strip(),
                            roll_no=roll_no,
                        )
                    )

        workbook.close()
        return students

    def _load_timetable(self) -> dict[tuple[str, int], list[dict]]:
        workbook = load_workbook(self.timetable_file, data_only=True)
        worksheet = workbook.active
        timetable: dict[tuple[str, int], list[dict]] = {}
        all_btech_depts = ["CE", "ME", "EE", "EC", "CS", "CH", "EL"]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[5] is None:
                continue

            program = str(row[0]).strip()
            sem_str = str(row[1]).strip().upper().replace("S", "")
            if not sem_str.isdigit():
                continue
            semester = int(sem_str)

            branch_raw = str(row[5]).strip().upper()
            if program == "B Arch":
                target_depts = ["B.ARCH"]
            elif branch_raw == "ALL BRANCHES":
                target_depts = all_btech_depts
            elif "/" in branch_raw:
                target_depts = [b.strip() for b in branch_raw.split("/")]
            else:
                target_depts = [branch_raw]

            exam_data = {
                "exam_date": str(row[2]).strip() if row[2] else "",
                "session": str(row[3]).strip() if row[3] else "",
                "subject_name": str(row[6]).strip() if row[6] else "",
                "subject_code": str(row[7]).strip() if row[7] else "",
            }

            for dept in target_depts:
                key = (dept, semester)
                if key not in timetable:
                    timetable[key] = []
                timetable[key].append(exam_data)

        workbook.close()
        return timetable

    @staticmethod
    def _get_department(sheet_name: str) -> str:
        raw = sheet_name.split()[0].upper().replace(".", "")
        if raw == "BARCH":
            return "B.ARCH"
        if raw == "EEE":
            return "EE"
        return raw

    @staticmethod
    def _get_semester(sheet_name: str) -> int:
        start = sheet_name.find("(S")
        end = sheet_name.find(")", start)
        return int(sheet_name[start + 2:end])

    @staticmethod
    def _get_section(sheet_name: str) -> str:
        parts = sheet_name.split()
        if len(parts) >= 3 and not parts[-2].startswith("(S"):
            return parts[-2]
        return ""