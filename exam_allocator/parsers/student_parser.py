"""
Excel student-list parser.

Reads the Students.xlsx workbook used by the exam seat allocation demo.

The workbook contains:
    - "Master Overview" sheet
    - One sheet per class

Each class sheet contains:
    - Class name
    - Semester
    - Duration
    - Student count information
    - Student roster

This module only parses and normalizes the Excel file.
It does not access the Django database.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
import re

from openpyxl import load_workbook


class StudentExcelParseError(Exception):
    """Raised when the student Excel workbook cannot be parsed."""


@dataclass
class StudentRecord:
    sl_no: str
    admission_no: str
    roll_number: str
    uni_reg_no: str
    name: str
    gender: str
    class_name: str
    semester: int
    academic_year: str


@dataclass
class ClassRecord:
    class_name: str
    semester: int
    academic_year: str
    department_name: str
    students: list[StudentRecord] = field(default_factory=list)


@dataclass
class StudentExtractionResult:
    source_file: str
    classes: list[ClassRecord]
    students: list[StudentRecord]
    issues: list[str] = field(default_factory=list)


ROSTER_HEADER = [
    "Sl. No.",
    "Admission No",
    "Roll No",
    "Uni Reg No",
    "Student Name",
    "Gender",
]

MASTER_SHEET_NAME = "Master Overview"


def parse_student_excel(excel_path: str) -> StudentExtractionResult:
    """
    Parse a Students.xlsx workbook.

    Returns normalized class and student records.
    """

    path = Path(excel_path)

    if not path.exists():
        raise StudentExcelParseError(
            f"File not found: {excel_path}"
        )

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise StudentExcelParseError(
            f"Unsupported file type: {path.suffix}"
        )

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise StudentExcelParseError(
            f"Could not open Excel workbook: {exc}"
        ) from exc

    classes: list[ClassRecord] = []
    students: list[StudentRecord] = []
    issues: list[str] = []

    try:
        for worksheet in workbook.worksheets:

            # The overview is useful for humans but the individual
            # class sheets contain the actual student records.
            if worksheet.title.strip() == MASTER_SHEET_NAME:
                continue

            try:
                class_record, class_students, class_issues = (
                    _parse_class_sheet(worksheet)
                )

                classes.append(class_record)
                students.extend(class_students)
                issues.extend(class_issues)

            except StudentExcelParseError as exc:
                issues.append(
                    f"Sheet '{worksheet.title}': {exc}"
                )

    finally:
        workbook.close()

    if not classes:
        raise StudentExcelParseError(
            "No class sheets containing student data were found."
        )

    return StudentExtractionResult(
        source_file=path.name,
        classes=classes,
        students=students,
        issues=issues,
    )


def _parse_class_sheet(worksheet):
    """
    Parse one class worksheet.
    """

    rows = list(
        worksheet.iter_rows(
            values_only=True
        )
    )

    if not rows:
        raise StudentExcelParseError(
            "Sheet is empty."
        )

    class_name = _extract_class_name(rows)

    semester = _extract_semester(rows)

    academic_year = _extract_academic_year(rows)

    department_name = _extract_department_name(rows)

    header_index = _find_roster_header(rows)

    if header_index is None:
        raise StudentExcelParseError(
            "Could not find student roster header."
        )

    student_rows = rows[header_index + 1 :]

    students, issues = _parse_students(
        student_rows=student_rows,
        class_name=class_name,
        semester=semester,
        academic_year=academic_year,
    )

    class_record = ClassRecord(
        class_name=class_name,
        semester=semester,
        academic_year=academic_year,
        department_name=department_name,
        students=students,
    )

    return class_record, students, issues


def _extract_class_name(rows) -> str:
    """
    Example source:

        Class Name: B.Arch 2K21A
    """

    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue

            text = str(cell).strip()

            match = re.search(
                r"Class Name\s*:\s*(.+)",
                text,
                re.IGNORECASE,
            )

            if match:
                return match.group(1).strip()

    raise StudentExcelParseError(
        "Class Name was not found."
    )


def _extract_semester(rows) -> int:
    """
    Example source:

        Semester: S10 (Xth Semester)

    Returns:

        10
    """

    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue

            text = str(cell).strip()

            match = re.search(
                r"Semester\s*:\s*S(\d+)",
                text,
                re.IGNORECASE,
            )

            if match:
                return int(match.group(1))

    raise StudentExcelParseError(
        "Semester was not found."
    )


def _extract_academic_year(rows) -> str:
    """
    Example source:

        Duration: 2021-2026

    The duration is used as the academic-year value for now.
    """

    for row in rows[:10]:
        for cell in row:
            if cell is None:
                continue

            text = str(cell).strip()

            match = re.search(
                r"Duration\s*:\s*(\d{4}\s*-\s*\d{4})",
                text,
                re.IGNORECASE,
            )

            if match:
                return match.group(1).replace(" ", "")

    return ""


def _extract_department_name(rows) -> str:
    """
    Example first row:

        TKM COLLEGE OF ENGINEERING, KOLLAM-5
        DEPARTMENT OF ARCHITECTURE
    """

    for row in rows[:3]:
        for cell in row:
            if cell is None:
                continue

            text = str(cell).strip()

            match = re.search(
                r"DEPARTMENT OF\s+(.+)",
                text,
                re.IGNORECASE,
            )

            if match:
                return (
                    "DEPARTMENT OF "
                    + match.group(1).strip()
                )

    return ""


def _find_roster_header(rows):
    """
    Find the row containing:

        Sl. No.
        Admission No
        Roll No
        Uni Reg No
        Student Name
        Gender
    """

    expected = [
        value.lower()
        for value in ROSTER_HEADER
    ]

    for index, row in enumerate(rows):

        values = [
            str(cell).strip().lower()
            if cell is not None
            else ""
            for cell in row[:6]
        ]

        if values == expected:
            return index

    return None


def _parse_students(
    student_rows,
    class_name: str,
    semester: int,
    academic_year: str,
):
    students: list[StudentRecord] = []
    issues: list[str] = []

    seen_roll_numbers: dict[str, int] = {}

    for row_number, row in enumerate(
        student_rows,
        start=1,
    ):

        values = list(row[:6])

        # Ignore completely empty rows.
        if not any(
            value is not None and str(value).strip()
            for value in values
        ):
            continue

        if len(values) < 6:
            issues.append(
                f"{class_name}: row {row_number} "
                f"has fewer than 6 columns. Skipped."
            )
            continue

        (
            sl_no,
            admission_no,
            roll_number,
            uni_reg_no,
            student_name,
            gender,
        ) = values

        sl_no = _clean_value(sl_no)
        admission_no = _clean_value(admission_no)
        roll_number = _clean_value(roll_number)
        uni_reg_no = _clean_value(uni_reg_no)
        student_name = _clean_value(student_name)
        gender = _clean_value(gender)

        # Stop if Excel contains another section/header.
        if roll_number.lower() == "roll no":
            continue

        if not roll_number:
            issues.append(
                f"{class_name}: row {row_number} "
                "has no roll number."
            )

        if not student_name:
            issues.append(
                f"{class_name}: row {row_number} "
                "has no student name."
            )

        if roll_number:

            if roll_number in seen_roll_numbers:
                issues.append(
                    f"{class_name}: duplicate roll number "
                    f"'{roll_number}' at row {row_number}; "
                    f"first seen at row "
                    f"{seen_roll_numbers[roll_number]}."
                )
            else:
                seen_roll_numbers[roll_number] = row_number

        student = StudentRecord(
            sl_no=sl_no,
            admission_no=admission_no,
            roll_number=roll_number,
            uni_reg_no=uni_reg_no,
            name=student_name,
            gender=gender,
            class_name=class_name,
            semester=semester,
            academic_year=academic_year,
        )

        students.append(student)

    return students, issues


def _clean_value(value) -> str:
    """
    Convert Excel values to clean strings.

    This is deliberately conservative so we don't accidentally
    modify roll numbers or names.
    """

    if value is None:
        return ""

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

    return str(value).strip()