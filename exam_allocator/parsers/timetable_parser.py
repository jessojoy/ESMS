"""
Excel timetable parser.

Reads Timetable.xlsx and converts each examination timetable
entry into a normalized Python object.

This module does not access Django models or the database.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook


class TimetableExcelParseError(Exception):
    """Raised when the timetable Excel workbook cannot be parsed."""


@dataclass
class ExamRecord:
    program: str
    semester: int
    exam_date: date
    session: str
    time: str
    branch: str
    branches: list[str]
    subject_name: str
    subject_code: str
    duration_minutes: int


@dataclass
class TimetableExtractionResult:
    source_file: str
    exams: list[ExamRecord]
    issues: list[str]


HEADER_NAMES = {
    "program",
    "semester",
    "date",
    "session",
    "time",
    "slot / branch",
    "subject name",
    "subject code",
}


def parse_timetable_excel(
    excel_path: str,
) -> TimetableExtractionResult:
    """
    Parse Timetable.xlsx.

    Returns normalized exam records.
    """

    path = Path(excel_path)

    if not path.exists():
        raise TimetableExcelParseError(f"File not found: {excel_path}")

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise TimetableExcelParseError(f"Unsupported file type: {path.suffix}")

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise TimetableExcelParseError(f"Could not open Excel workbook: {exc}") from exc

    exams: list[ExamRecord] = []
    issues: list[str] = []

    try:
        for worksheet in workbook.worksheets:

            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                continue

            header_index = _find_header_row(rows)

            if header_index is None:
                issues.append(
                    f"Sheet '{worksheet.title}': " "timetable header not found."
                )
                continue

            headers = _normalize_headers(rows[header_index])

            for row_number, row in enumerate(
                rows[header_index + 1 :],
                start=header_index + 2,
            ):

                if _is_empty_row(row):
                    continue

                try:
                    exam = _parse_exam_row(
                        row=row,
                        headers=headers,
                    )

                    if exam is not None:
                        exams.append(exam)

                except TimetableExcelParseError as exc:
                    issues.append(
                        f"Sheet '{worksheet.title}', " f"row {row_number}: {exc}"
                    )

    finally:
        workbook.close()

    if not exams:
        raise TimetableExcelParseError("No timetable entries were found.")

    return TimetableExtractionResult(
        source_file=path.name,
        exams=exams,
        issues=issues,
    )


def _find_header_row(rows):
    """
    Find the timetable header row.
    """

    for index, row in enumerate(rows):

        normalized = {_normalize_header(cell) for cell in row if cell is not None}

        matches = len(normalized.intersection(HEADER_NAMES))

        if matches >= 5:
            return index

    return None


def _normalize_headers(row) -> list[str]:
    """
    Normalize header names while retaining their positions.
    """

    return [_normalize_header(cell) for cell in row]


def _normalize_header(value) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()

    text = re.sub(
        r"\s+",
        " ",
        text,
    )

    return text


def _parse_exam_row(row, headers) -> ExamRecord | None:

    values = {}

    for index, header in enumerate(headers):

        if not header:
            continue

        if index >= len(row):
            values[header] = None
        else:
            values[header] = row[index]

    program = _clean_text(values.get("program"))

    semester = _parse_semester(values.get("semester"))

    exam_date = _parse_date(values.get("date"))

    session = _parse_session(values.get("session"))

    time = _clean_text(values.get("time"))

    branch = _clean_text(values.get("slot / branch"))

    subject_name = _clean_text(values.get("subject name"))

    subject_code = _clean_text(values.get("subject code"))

    # Your requested fallback:
    # if there is no subject code, use subject name.
    if not subject_code or subject_code.upper() in {
        "N/A",
        "NA",
        "NONE",
        "NULL",
        "-",
    }:

        subject_code = subject_name

    if not program:
        raise TimetableExcelParseError("Program is missing.")

    if semester is None:
        raise TimetableExcelParseError("Semester is missing.")

    if exam_date is None:
        raise TimetableExcelParseError("Date is missing.")

    if not session:
        raise TimetableExcelParseError("Session is missing.")

    if not branch:
        raise TimetableExcelParseError("Branch is missing.")

    if not subject_name:
        raise TimetableExcelParseError("Subject name is missing.")

    if not subject_code:
        raise TimetableExcelParseError(
            "Subject code and subject name are both missing."
        )

    branches = _parse_branches(branch)

    duration_minutes = _calculate_duration_minutes(time)

    return ExamRecord(
        program=program,
        semester=semester,
        exam_date=exam_date,
        session=session,
        time=time,
        branch=branch,
        branches=branches,
        subject_name=subject_name,
        subject_code=subject_code,
        duration_minutes=duration_minutes,
    )


def _parse_semester(value) -> int | None:

    if value is None:
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)

    text = str(value).strip()

    match = re.search(
        r"S?\s*(\d+)",
        text,
        re.IGNORECASE,
    )

    if not match:
        return None

    return int(match.group(1))


def _parse_date(value) -> date | None:

    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    # Examples:
    # 12-03-2026 (Thursday)
    # 12/03/2026
    # 12-03-2026

    match = re.search(
        r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})",
        text,
    )

    if not match:
        return None

    day = int(match.group(1))
    month = int(match.group(2))
    year = int(match.group(3))

    try:
        return date(
            year,
            month,
            day,
        )
    except ValueError:
        return None


def _parse_session(value) -> str:

    if value is None:
        return ""

    text = str(value).strip().upper()

    if text.startswith("FN"):
        return "FN"

    if text.startswith("AN"):
        return "AN"

    return text


def _parse_branches(value) -> list[str]:

    if not value:
        return []

    text = str(value).strip()

    # Normalize common separators.
    text = text.replace(",", "/")
    text = text.replace("&", "/")
    text = text.replace("+", "/")

    branches = [item.strip() for item in text.split("/") if item.strip()]

    return branches


def _calculate_duration_minutes(
    time_text: str,
) -> int:
    """
    Convert common timetable ranges into duration minutes.

    Example:
        10:00 - 12:00 Noon -> 120

    If the time cannot be interpreted, return 0.
    """

    if not time_text:
        return 0

    match = re.search(
        r"(\d{1,2}):(\d{2})\s*[-–]\s*" r"(\d{1,2}):(\d{2})",
        time_text,
    )

    if not match:
        return 0

    start_hour = int(match.group(1))
    start_minute = int(match.group(2))

    end_hour = int(match.group(3))
    end_minute = int(match.group(4))

    start_total = start_hour * 60 + start_minute

    end_total = end_hour * 60 + end_minute

    duration = end_total - start_total

    if duration < 0:
        duration += 24 * 60

    return duration


def _clean_text(value) -> str:

    if value is None:
        return ""

    return re.sub(
        r"\s+",
        " ",
        str(value).strip(),
    )


def _is_empty_row(row) -> bool:

    return not any(value is not None and str(value).strip() for value in row)
