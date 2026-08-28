"""
Excel classroom parser.

Reads Class.xlsx and converts classroom information into
normalized Python objects.

The current Class.xlsx contains classroom numbers.
Capacity, benches and building are supplied through configurable
defaults because they are not present in the source workbook.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


class ClassroomExcelParseError(Exception):
    """Raised when the classroom Excel workbook cannot be parsed."""


@dataclass
class ClassroomRecord:
    room_number: str
    capacity: int
    benches: int
    building: str


@dataclass
class ClassroomExtractionResult:
    source_file: str
    classrooms: list[ClassroomRecord]
    issues: list[str]


DEFAULT_CAPACITY = 45
DEFAULT_BENCHES = 15
DEFAULT_BUILDING = "Main Building"


def parse_classroom_excel(
    excel_path: str,
    default_capacity: int = DEFAULT_CAPACITY,
    default_benches: int = DEFAULT_BENCHES,
    default_building: str = DEFAULT_BUILDING,
) -> ClassroomExtractionResult:
    """
    Parse Class.xlsx.

    Parameters
    ----------
    excel_path:
        Path to the classroom Excel workbook.

    default_capacity:
        Capacity assigned to rooms when the workbook does not
        provide one.

    default_benches:
        Number of benches assigned to rooms when the workbook
        does not provide one.

    default_building:
        Building assigned to rooms when the workbook does not
        provide one.
    """

    path = Path(excel_path)

    if not path.exists():
        raise ClassroomExcelParseError(
            f"File not found: {excel_path}"
        )

    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ClassroomExcelParseError(
            f"Unsupported file type: {path.suffix}"
        )

    if default_capacity <= 0:
        raise ClassroomExcelParseError(
            "Default capacity must be greater than zero."
        )

    if default_benches <= 0:
        raise ClassroomExcelParseError(
            "Default benches must be greater than zero."
        )

    try:
        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ClassroomExcelParseError(
            f"Could not open Excel workbook: {exc}"
        ) from exc

    classrooms: list[ClassroomRecord] = []
    issues: list[str] = []

    try:
        for worksheet in workbook.worksheets:

            for row_number, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):

                if not row:
                    continue

                room_value = row[0]

                if room_value is None:
                    continue

                room_number = str(room_value).strip()

                if not room_number:
                    continue

                # Skip the header row.
                if room_number.lower() in {
                    "class room",
                    "classroom",
                    "room",
                    "room number",
                    "room no",
                }:
                    continue

                # Avoid accidental duplicate rooms.
                if any(
                    classroom.room_number == room_number
                    for classroom in classrooms
                ):
                    issues.append(
                        f"Duplicate classroom "
                        f"'{room_number}' found at "
                        f"row {row_number}. Skipped."
                    )
                    continue

                classrooms.append(
                    ClassroomRecord(
                        room_number=room_number,
                        capacity=default_capacity,
                        benches=default_benches,
                        building=default_building,
                    )
                )

    finally:
        workbook.close()

    if not classrooms:
        raise ClassroomExcelParseError(
            "No classrooms were found in the workbook."
        )

    return ClassroomExtractionResult(
        source_file=path.name,
        classrooms=classrooms,
        issues=issues,
    )